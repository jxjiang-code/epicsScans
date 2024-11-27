import sys
import numpy as np
import pandas as pd
import openpyxl
import json
import epics
import time
import PyQt5.QtCore as Qt
from PyQt5.QtGui import QColor
from scipy.optimize import curve_fit
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QLabel, QLineEdit, QPushButton, QComboBox, QTableWidget, QTableWidgetItem, QHBoxLayout, QCheckBox, QMenuBar, QAction, QFileDialog, QMessageBox, QDialog)
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from PyQt5.QtCore import QTimer

# Define the error function for Z alignment
def error_function(x, x0, scale, width):
    return scale * (1 + np.tanh((x - x0) / width)) / 2


# Define the Gaussian function for Theta alignment
def gaussian(x, x0, width, scale):
    return scale * np.exp(-((x - x0) ** 2) / (2 * width ** 2))

# Define fitting functions
def linear(x, slope, intercept):
    return slope * x + intercept

def lorentz(x, a, x0, gamma):
    return a / (1 + ((x - x0) / gamma) ** 2)


class DynamicPlot(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Dynamic Plot with One Dimensional Scan")
        self.setGeometry(100, 100, 800, 800)

        # Create main widget
        self.main_widget = QWidget(self)
        self.setCentralWidget(self.main_widget)

        # Create layout
        layout = QVBoxLayout(self.main_widget)


        # Add a menu bar
        self.create_menu_bar()

        # Label for the dropdown
        self.msglabel1 = QLabel(f"<span style='font-size:16pt; font-weight:bold; color:blue;'>Press scan bottom to start a scanning ...")
        layout.addWidget(self.msglabel1)

        # Create Matplotlib figure and canvas
        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)

        # Initialize crosshair points and artists
        self.left_cross = None  # Position of left crosshair
        self.right_cross = None  # Position of right crosshair
        self.left_marker = None  # Matplotlib artist for left crosshair
        self.right_marker = None  # Matplotlib artist for right crosshair
        self.middle_marker = None  # Matplotlib artist for the middle point

        # Connect the event handler
        self.cid = self.canvas.mpl_connect("button_press_event", self.on_click)

        # Initialize plot
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title("Dynamic Data")
        self.ax.set_xlabel("X-axis")
        self.ax.set_ylabel("Y-axis")

        # Timer for updating random data
        #self.timer = QTimer(self)
        #self.timer.timeout.connect(self.update_live_plot)
        #self.timer.start(1000)  # Update every 1 second

        # Data for the plot
        self.data = {"x": [], "y": [], "label": "Live Data"}

        # Create a horizontal layout for checkboxes
        scanning_layout = QHBoxLayout()

        # Label for the dropdown
        self.label1 = QLabel("Select a motor from the dropdown:")
        scanning_layout.addWidget(self.label1)

        # Create the dropdown (QComboBox)
        self.dropdown1 = QComboBox()
        scanning_layout.addWidget(self.dropdown1)

        # Label for the dropdown
        self.label2 = QLabel("Select a detector from the dropdown:")
        scanning_layout.addWidget(self.label2)

        # Create the dropdown (QComboBox)
        self.dropdown2 = QComboBox()
        scanning_layout.addWidget(self.dropdown2)

        # Load data into the dropdown
        self.load_excel_data()

        # Connect dropdown selection change to a method
        self.dropdown1.currentTextChanged.connect(self.on_dropdown1_change)
        self.dropdown2.currentTextChanged.connect(self.on_dropdown2_change)

        # Add the horizontal box into the layout
        layout.addLayout(scanning_layout)
        
        # Add label and textbox for optimized values
        # Create a horizontal layout for checkboxes
        optimize_layout = QHBoxLayout()
        # Label for the dropdown
        self.optimized_label = QLabel("Optimized %s:" % self.dropdown1.itemText(0))
        self.text = {"motor": self.dropdown1.itemText(0), "detector": self.dropdown2.itemText(0)}
        self.data["label"] = self.dropdown1.itemText(0)
        self.optimized_input = QLineEdit()
        optimize_layout.addWidget(self.optimized_label)
        optimize_layout.addWidget(self.optimized_input)


        # Add checkboxes for fitting options
        self.add_fitting_options(layout)

        # Add label and textbox for optimized values
        layout.addLayout(optimize_layout)
        
        # Mapping checkboxes to functions
        self.function_map = {
            "Linear": linear,
            "Gaussian": gaussian,
            "Lorentz": lorentz,
            "Error function": error_function,
        }


        # Add table for scan parameters
        self.add_scan_parameters_table(layout)

        # Buttons for alignment
        self.scan_button = QPushButton("SCAN")
        self.scan_button.clicked.connect(self.scan)
        # Change button color to blue and text to white
        self.scan_button.setStyleSheet("""
            QPushButton {
                background-color: grey; 
                color: white;
                font-weight: bold;  /* Optional: Make text bold */
                border-radius: 5px; /* Optional: Add rounded corners */
                padding: 5px;       /* Optional: Add padding for a nicer look */
            }
            QPushButton:hover {
                background-color: lightblue; /* Optional: Change color on hover */
            }
        """)

        layout.addWidget(self.scan_button)

    def add_scan_parameters_table(self, layout):
        # Create a table with 2 rows and 6 columns
        self.table = QTableWidget(2, 6)
        self.table.setHorizontalHeaderLabels(["Start", "Middle", "End", "Step", "Num. of Points", "Time"])
        self.table.setVerticalHeaderLabels(["Parameter 1", "Parameter 2"])

        # Initialize table with default values
        default_values = [
            ["0.000",  "5.000", "10.000", "1.000", "11", "0.500"],
            ["0.000", "10.000", "20.000", "2.000", "11", "1.000"]
        ]
        for row in range(2):
            for col in range(6):
                item = QTableWidgetItem(default_values[row][col])
                self.table.setItem(row, col, item)

        # Connect itemChanged to update dependent values
        self.table.itemChanged.connect(self.on_table_item_changed)

        # Add table to layout
        layout.addWidget(self.table)

    def on_click(self, event):
        """Handle mouse click events."""
        if event.inaxes != self.ax:
            return  # Ignore clicks outside the plot area

        if event.button == 1:  # Left click
            self.left_cross = (event.xdata, event.ydata)
            # Update or create the left crosshair marker
            if self.left_marker:
                self.left_marker.set_data(event.xdata, event.ydata)
            else:
                self.left_marker, = self.ax.plot(
                    event.xdata, event.ydata, "rx", label="Left Crosshair"
                )
            #print(f"Left crosshair: {self.left_cross}")

        elif event.button == 3:  # Right click
            self.right_cross = (event.xdata, event.ydata)
            # Update or create the right crosshair marker
            if self.right_marker:
                self.right_marker.set_data(event.xdata, event.ydata)
            else:
                self.right_marker, = self.ax.plot(
                    event.xdata, event.ydata, "bo", label="Right Crosshair"
                )
            #print(f"Right crosshair: {self.right_cross}")

        # Recalculate middle position and average if both crosshairs are placed
        if self.left_cross and self.right_cross:
            middle_position = (
                (self.left_cross[0] + self.right_cross[0]) / 2,
                (self.left_cross[1] + self.right_cross[1]) / 2,
            )
            average_value = (self.left_cross[1] + self.right_cross[1]) / 2
            #print(f"Middle position: {middle_position}")
            #print(f"Average value: {average_value}")

            # Update or create the middle marker
            if self.middle_marker:
                self.middle_marker.set_data(middle_position[0], middle_position[1])
            else:
                self.middle_marker, = self.ax.plot(
                    middle_position[0],
                    middle_position[1],
                    "g+",
                    label="Middle Position",
                )

        if self.left_cross != None and self.right_cross != None:
            self.msglabel1.setText(
        f"<span style='font-size:16pt; font-weight:bold; color:blue;'>Left crosshair:</span> (%.02f, %.02f), "
        f"<span style='font-size:16pt; font-weight:bold; color:green;'>Middle position:</span> (%.02f, %.02f), "  
        f"<span style='font-size:16pt; font-weight:bold; color:red;'>Right crosshair:</span> (%.02f, %.02f) " %
        (*self.left_cross, *middle_position, *self.right_cross))
        elif self.left_cross != None:
            self.msglabel1.setText(f"<span style='font-size:16pt; font-weight:bold; color:blue;'>Left crosshair:</span> (%.02f, %.02f)" % self.left_cross)
        elif self.right_cross != None:
            self.msglabel1.setText(f"<span style='font-size:16pt; font-weight:bold; color:red;'>Right crosshair:</span> (%.02f, %.02f)" % self.right_cross)
        self.ax.legend()
        self.canvas.draw()


    def on_table_item_changed(self, item):
        """Recalculate dependent values when Start, End, or Step changes."""
        # Block signals to avoid recursive triggering
        self.table.blockSignals(True)
        try:
            row = item.row()
            col = item.column()
            
            # Read the current values for calculation
            start = float(self.table.item(row, 0).text())
            middle = float(self.table.item(row, 1).text())
            end = float(self.table.item(row, 2).text())
            step = float(self.table.item(row, 3).text())
            num = int(float(self.table.item(row, 4).text()))

            
            #print(row, col, start, middle, end, step, num)
            if (col == 0) or (col == 2):  # Only recalculate scan parameters
                # Recalculate Middle
                middle = (start + end) / 2
                self.table.item(row, 0).setText(f"{start:.3f}")
                self.table.item(row, 1).setText(f"{middle:.3f}")
                self.table.item(row, 2).setText(f"{end:.3f}")
                # Recalculate step`
                step = (end - start)/(num-1)
                self.table.item(row, 3).setText(f"{step:.3f}")
            elif col == 1:
                # Recalculate End
                end =  2*middle - start
                self.table.item(row, 1).setText(f"{middle:.3f}")
                self.table.item(row, 2).setText(f"{end:.3f}")
                # Recalculate step`
                step = (end - start)/(num-1)
                self.table.item(row, 3).setText(f"{step:.3f}")
            elif col == 3:
                # Recalculate Number`
                if step != 0:
                    num = abs(int((end - start)/step)) + 1
                    if (end - start)/step < 0:
                        self.table.item(row, 0).setText(f"{end:.3f}")
                        self.table.item(row, 2).setText(f"{start:.3f}")
                    self.table.item(row, 3).setText(f"{step:.3f}")
                    self.table.item(row, 4).setText(f"{num}")
                else:
                    self.table.item(row, 3).setText(f"{step:.3f}")
                    self.table.item(row, 4).setText("0")
            elif col == 4:
                if num > 1:
                    # Recalculate step`
                    step = (end - start)/(num-1)
                    self.table.item(row, 3).setText(f"{step:.3f}")
                    self.table.item(row, 4).setText(f"{num}")
                else:
                    step = end - start
                    self.table.item(row, 3).setText(f"{step:.3f}")
                    self.table.item(row, 4).setText("1")
        finally:
            # Re-enable signals after programmatic updates
            self.table.blockSignals(False) 
            

    def add_fitting_options(self, layout):
        # Create a horizontal layout for checkboxes
        fitting_layout = QHBoxLayout()
        self.checkboxes = {}

        # Label for the dropdown
        self.fitting_label = QLabel("Select a fitting function:      ")
        fitting_layout.addWidget(self.fitting_label)


        # Add checkboxes for each fitting function
        for fit_type in ["Linear", "Gaussian", "Lorentz", "Error function"]:
            checkbox = QCheckBox(fit_type)
            self.checkboxes[fit_type] = checkbox
            fitting_layout.addWidget(checkbox)

        layout.addLayout(fitting_layout)


    def load_excel_data(self):
        """Load data from an Excel file into the dropdown."""
        try:
            # Load the Excel file
            file_path = "./scan_pvs_table.xlsx"  # Replace with your Excel file path
            self.pvList = pd.read_excel(file_path)
            print(self.pvList)

            # Check if the column 'alias' exists
            if "Alias" in self.pvList.columns:
                # Populate the dropdown with unique values from the 'alias' column
                self.dropdown1.addItems(self.pvList[self.pvList["Type"]=="Motor"]["Alias"].dropna().unique())
                self.dropdown2.addItems(self.pvList[self.pvList["Type"]=="Detector"]["Alias"].dropna().unique())
            else:
                self.label1.setText("Column 'alias' not found in the Excel file.")
                self.label2.setText("Column 'alias' not found in the Excel file.")
        except Exception as e:
            self.label1.setText(f"Error loading Excel file: {e}")
            self.label2.setText(f"Error loading Excel file: {e}")

    def on_dropdown1_change(self, text):
        """Handle the dropdown selection change."""
        self.label1.setText(f"Selected motor: {text}")
        self.optimized_label.setText(f"Optimized {text}:" )
        self.text["motor"] = text
    def on_dropdown2_change(self, text):
        self.label2.setText(f"Selected detector: {text}")
        self.text["detector"] = text

    def update_plot(self):
        """Update the plot with current data."""
        self.ax.clear()
        self.ax.set_title(self.data["label"])
        self.ax.set_xlabel("%s (%s)"  % (self.text['motor'], self.pvList.loc[self.pvList["Alias"] == self.text["motor"], "EGU"].values[0]))
        self.ax.set_ylabel("%s (%s)"  % (self.text['detector'], self.pvList.loc[self.pvList["Alias"] == self.text["detector"], "EGU"].values[0]))
        self.ax.plot(self.data["x"], self.data["y"], "o-", label="Data")
        for n in np.arange(len(self.checked_names)):
            checked_name = self.checked_names[n]
            self.ax.plot(self.data[checked_name]["fit_x"], self.data[checked_name]["fit_y"], "-", label="%s Fitting" % checked_name)
        self.ax.legend()
        self.canvas.draw()

    def scan(self):

        # Display the message as scanning 
        self.msglabel1.setText(f"<span style='font-size:16pt; font-weight:bold; color:green;'>Scanning ...")

        # Re-Initialize crosshair points and artists
        self.left_cross = None  # Position of left crosshair
        self.right_cross = None  # Position of right crosshair
        self.left_marker = None  # Matplotlib artist for left crosshair
        self.right_marker = None  # Matplotlib artist for right crosshair
        self.middle_marker = None  # Matplotlib artist for the middle point

        """Perform motor scan, record detector and update the plot."""
        # Timer for updating random data
        self.current_index = 0
        # Data for the plot
        self.data = {"x": [], "y": [], "label": self.text["motor"]}

        """List all checked checkboxes."""
        self.checked_names = [name for name, checkbox in self.checkboxes.items() if checkbox.isChecked()]
        for n in np.arange(len(self.checked_names)):
            self.data[self.checked_names[n]] = {"fit_x": [], "fit_y": [], "optimized values":[]}
        
        # Read the current values for calculation
        row = 0
        self.start = float(self.table.item(row, 0).text())
        self.middle = float(self.table.item(row, 1).text())
        self.end = float(self.table.item(row, 2).text())
        self.step = float(self.table.item(row, 3).text())
        self.num = int(float(self.table.item(row, 4).text()))
        self.accu = float(self.table.item(row, 5).text())
        self.scanPos = np.linspace(self.start, self.end, self.num)

        self.scan_timer = QTimer(self)
        self.scan_timer.timeout.connect(self.update_scan_step)
        self.scan_timer.start(500)  # Update every 1 second

    def update_scan_step(self):
        """Update the plot for each step during 1d scan."""

        # Incrementally process data
        i = self.current_index
        if i >= self.num:
            self.scan_timer.stop()  # Stop the timer when all data is processed

            # Display the message as scanning 
            self.msglabel1.setText(f"<span style='font-size:16pt; font-weight:bold; color:green;'>Scan finished!")

            # Display optimized value
            self.optimized_input.setText("%s" % [[self.checked_names[n], self.data[self.checked_names[n]]['optimized values']] for n in np.arange(len(self.checked_names))])

            return

        try:
            # Update the data to the current step
            motor_pv = self.pvList.loc[self.pvList["Alias"] == self.text["motor"], "PV"].values[0]+".VAL"
            epics.caput(motor_pv, self.scanPos[i], timeout=4) 
            time.sleep(self.accu * 0.1)  # Allow time for movement

            # Read the motor position
            motor_rbv = motor_pv.replace(".VAL", ".RBV")
            motor_position = epics.caget(motor_rbv, timeout=4)
            if motor_position is None:
                raise TimeoutError(f"Timeout reading motor position from {motor_rbv}")
            self.data["x"].append(motor_position)

            # Count  detector for accumulate time
            time.sleep(self.accu)

            detector_pv = self.pvList.loc[self.pvList["Alias"] == self.text["detector"], "PV"].values[0]
            detector_value = epics.caget(detector_pv, timeout=4)
            if detector_value is None:
                raise TimeoutError(f"Timeout reading detector value from {detector_pv}")
            self.data["y"].append(detector_value)

            if i > 1:
                # Fit to any function for the current data slice
                try:
                    for n in np.arange(len(self.checked_names)):
                        checked_name = self.checked_names[n]
                        popt, _ = curve_fit(self.function_map[checked_name], self.data['x'], self.data['y'])
                        #popt, _ = curve_fit(self.function_map[checked_name], self.data['x'], self.data['y'], p0=[0, 0.1, 1])
                        self.data[checked_name]["optimized values"] = list(popt)  # Save the best fitting value
                        self.x_fine_values = np.linspace(self.data['x'][0], self.data['x'][-1], len(self.data['x']) * 10)
                        self.data[checked_name]["fit_x"], self.data[checked_name]["fit_y"] = list(self.x_fine_values), list(self.function_map[checked_name](self.x_fine_values, *popt))
                except RuntimeError:
                    pass  # Ignore fitting errors for small data points
            else:
                self.data["fit_x"], self.data["fit_y"] = [], []
            
            # Save the PV and EGU information
            self.data["scan"], self.data["fitting"] = self.text, self.checked_names 

            # Update the plot
            self.update_plot()

            # Move to the next data point
            self.current_index += 1

            #print(self.data['x'])

        except (TimeoutError, epics.ca.ChannelAccessException) as e:

            # Handle timeout or EPICS communication errors

            self.msglabel1.setText(f"<span style='font-size:16pt; font-weight:bold; color:red;'>Error during scan step: {e}")
            self.scan_timer.stop()  # Stop the scan
            return 

    def create_menu_bar(self):
        """Create the menu bar with File and Edit options."""
        menu_bar = self.menuBar()

        # File menu
        file_menu = menu_bar.addMenu("File")

        # Add 'Load' action in File menu
        load_action = QAction('Load', self)
        load_action.triggered.connect(self.load_data)
        file_menu.addAction(load_action)

        save_action = QAction("Save", self)
        save_action.triggered.connect(self.save_data)
        file_menu.addAction(save_action)

        save_figure_action = QAction("Save the Figure", self)
        save_figure_action.triggered.connect(self.save_figure)
        file_menu.addAction(save_figure_action)

        # Edit menu
        edit_menu = menu_bar.addMenu("Edit")

        edit_excel_action = QAction("Edit the Excel File", self)
        edit_excel_action.triggered.connect(self.edit_excel_file)
        edit_menu.addAction(edit_excel_action)

    def save_data(self):
        """Save the data to a file."""
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Data", "", "CSV Files (*.json);;All Files (*)", options=options)
        if file_name:
            try:
                # Write the dictionary to the selected file in JSON format
                with open(file_name, 'w') as f:
                    json.dump(self.data, f, indent=4)
                QMessageBox.information(self, "Save Data", f"Data successfully saved to {file_name}")
            except Exception as e:
                QMessageBox.warning(self, "Save Data", f"Error saving data: {str(e)}")

    def save_figure(self):
        """Save the figure to a file."""
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Figure", "", "PNG Files (*.png);;All Files (*)", options=options)
        if file_name:
            self.figure.savefig(file_name)
            QMessageBox.information(self, "Save Figure", f"Figure saved to {file_name}")

    def load_data(self):
        """Load data from JSON and update the plot."""
        # Open file dialog to select the JSON file
        file_name, _ = QFileDialog.getOpenFileName(self, "Open JSON File", "", "JSON Files (*.json)")
        if file_name:
            try:
                # Load data from JSON file
                with open(file_name, 'r') as file:
                    loaded_data = json.load(file)
                
                # Update self.data with loaded data
                self.data = loaded_data

                # Update the plot with the loaded data
                self.load_plot()

            except Exception as e:
                print(f"Error loading data: {e}")

    def load_plot(self):
        """Update the plot with current data."""
        self.ax.clear()
        
        # Reload the json files to the variables
        self.text = self.data["scan"]
        self.checked_names = self.data["fitting"]

        self.ax.set_title(self.data["label"])
        self.ax.set_xlabel("%s (%s)"  % (self.text['motor'], self.pvList.loc[self.pvList["Alias"] == self.text["motor"], "EGU"].values[0]))
        self.ax.set_ylabel("%s (%s)"  % (self.text['detector'], self.pvList.loc[self.pvList["Alias"] == self.text["detector"], "EGU"].values[0]))
        self.ax.plot(self.data["x"], self.data["y"], "o-", label="Data")
        for n in np.arange(len(self.checked_names)):
            checked_name = self.checked_names[n]
            self.ax.plot(self.data[checked_name]["fit_x"], self.data[checked_name]["fit_y"], "-", label="%s Fitting" % checked_name)
        self.ax.legend()
        self.canvas.draw()

    def edit_excel_file(self):
        """Edit the Excel file."""
        # Open file dialog to select an Excel file
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx;*.xls);;All Files (*)", options=options)
        
        if file_name:
            try:
                # Load the Excel file using openpyxl
                wb = openpyxl.load_workbook(file_name)
                sheet = wb.active

                # Show a message box to confirm the file loading
                QMessageBox.information(self, "Edit Excel", f"Opened Excel file: {file_name}")
                
                # Show the dialog box for editing Excel content
                self.show_table_edit_dialog(wb, sheet, file_name)

            except Exception as e:
                QMessageBox.warning(self, "Error", f"Error opening the Excel file: {e}")

    def show_table_edit_dialog(self, wb, sheet, file_name):
        """Show a dialog with a table to edit the Excel content."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Excel Content")

        # Layout for dialog
        layout = QVBoxLayout()

        # Create table widget to display the data
        table_widget = QTableWidget()
        num_rows = sheet.max_row
        num_cols = sheet.max_column
        table_widget.setRowCount(num_rows)
        table_widget.setColumnCount(num_cols)

        # Fill the table with Excel data
        for row in range(num_rows):
            for col in range(num_cols):
                value = sheet.cell(row=row+1, column=col+1).value
                item = QTableWidgetItem(str(value))

                # Make the first row uneditable
                if row == 0:  
                    item.setFlags(item.flags() & ~Qt.Qt.ItemIsEditable)  
                    item.setBackground(QColor("lightgrey"))  

                table_widget.setItem(row, col, item)

        # Add Save button to save the edited content back to the Excel file
        save_button = QPushButton("Save")
        save_button.clicked.connect(lambda: self.save_table_data(table_widget, file_name))

        # Add Add Row button
        add_row_button = QPushButton("Add Row")
        add_row_button.clicked.connect(lambda: self.add_row(table_widget))

        # Add Delete Row button
        delete_row_button = QPushButton("Delete Row")
        delete_row_button.clicked.connect(lambda: self.delete_row(table_widget))

        # Add table widget, Add Row button, Delete Row button, and Save button to the layout
        layout.addWidget(table_widget)
        layout.addWidget(add_row_button)
        layout.addWidget(delete_row_button)
        layout.addWidget(save_button)

        dialog.setLayout(layout)

        dialog.exec_()

    def add_row(self, table_widget):
        """Add a new row to the table."""
        row_position = table_widget.rowCount()
        table_widget.insertRow(row_position)
        # Optionally, set the cells of the new row to empty values
        for col in range(table_widget.columnCount()):
            table_widget.setItem(row_position, col, QTableWidgetItem(''))

    def delete_row(self, table_widget):
        """Delete the selected row from the table."""
        selected_row = table_widget.currentRow()
        if selected_row >= 0:  # If a row is selected
            table_widget.removeRow(selected_row)
        else:
            QMessageBox.warning(self, "No Row Selected", "Please select a row to delete.")

    def save_table_data(self, table_widget, file_name):
        """Save the edited data from the table back to a new Excel file."""
        try:
            num_rows = table_widget.rowCount()
            num_cols = table_widget.columnCount()

            # Create a new workbook and sheet
            wb = openpyxl.Workbook()
            sheet = wb.active

            # Write the table data to the new workbook
            for row in range(num_rows):
                for col in range(num_cols):
                    item = table_widget.item(row, col)
                    value = item.text() if item else ""  # Ensure no errors for empty cells
                    sheet.cell(row=row + 1, column=col + 1, value=value)

            # Save the new workbook to the specified file
            wb.save(file_name)

            # Reload the dropdowns to reflect changes
            self.dropdown1.clear()
            self.dropdown2.clear()
            self.load_excel_data()

            QMessageBox.information(self, "Save Successful", f"Excel file saved successfully: {file_name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving the Excel file: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = DynamicPlot()
    main_window.show()
    sys.exit(app.exec_())

